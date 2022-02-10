#-*- coding: utf-8 -*-

from os import path, walk, getcwd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import sys

if sys.argv[1:]:
    self = ' '.join(sys.argv[1:])
else:
    self = getcwd()
    
# 폴더의 파일 읽기

REMOVES = ['makelist.exe', 'renamer.xlsx', 'renamer.exe']

try:
    walk = list(walk(self))
except:
    print('cannot read directory')
    quit()

for rem in REMOVES:
    try:
        walk[0][2].remove(rem)
    except:
        continue

items = []

for step in walk:
    for folder in step[1]:
        items.append([step[0]+"\\"+folder+"\\",""])
    for file in step[2]:
        items.append([step[0]+"\\", file])

# 엑셀 파일 생성

wb = Workbook()
ws = wb.active
ws.title = "rename"
ws.append(['','원래 경로', '원래 이름', '변경할 경로', '변경할 이름'])

for item in items:
    type = ['[파일]'] if path.isfile('\\'.join(item)) else ['[폴더]']
    ws.append(type+item+item)

# 시트 스타일링
c = PatternFill(fill_type='solid', start_color='F6F6F6', end_color='F6F6F6')
b = Font(bold=True)

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 3

for col in ws.iter_cols(min_col=1, max_col=3):
    for cell in col:
        cell.fill = c
for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = b

# 저장

try:
    wb.save('renamer.xlsx')
    for item in items:
        print(item)
except:
    print("error. check if renamer.xlsx is being used")